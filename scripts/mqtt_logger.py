"""
ESP32 MQTT Data Logger
Connects to test.enarxi.com:1883, subscribes to battery & solar voltage topics,
and logs every reading to esp32_data.xlsx.

Excel format:
    Timestamp | Battery Voltage | Solar Voltage

Usage:
    pip install paho-mqtt openpyxl
    python mqtt_logger.py
"""

import os
import sys
import signal
import datetime
import uuid
import paho.mqtt.client as mqtt
from openpyxl import Workbook, load_workbook

# -- Configuration -----------------------------------------------
BROKER = "test.enarxi.com"
PORT = 1883
TOPICS = [
    "esp32/solar/voltage",
    "esp32/battery/voltage",
]
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "esp32_data.xlsx")
IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))
# ----------------------------------------------------------------

message_count = 0
# Buffer to pair up battery + solar before writing one row
reading_buffer = {"battery": None, "solar": None}


def get_or_create_workbook(path):
    """Load existing workbook or create a new one with headers."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "ESP32 Data"
        ws.append(["Timestamp", "Battery Voltage", "Solar Voltage"])
        # Column widths for readability
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        wb.save(path)
    return wb, ws


def flush_row(userdata):
    """Write a combined row when both values are available, then reset buffer."""
    global message_count, reading_buffer

    now = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    battery = reading_buffer["battery"]
    solar = reading_buffer["solar"]

    ws, wb = userdata["ws"], userdata["wb"]
    ws.append([now, battery, solar])
    wb.save(EXCEL_FILE)

    message_count += 1
    print(f"[{now}]  Battery={battery}V  |  Solar={solar}V")

    # Reset buffer for next pair
    reading_buffer["battery"] = None
    reading_buffer["solar"] = None


# -- MQTT Callbacks ---------------------------------------------
def on_connect(client, userdata, flags, rc, properties=None):
    if rc == 0:
        print(f"[OK] Connected to {BROKER}:{PORT}")
        for t in TOPICS:
            client.subscribe(t)
            print(f"    Subscribed -> {t}")
    else:
        print(f"[FAIL] Connection failed (rc={rc})")
        sys.exit(1)


def on_message(client, userdata, msg):
    global reading_buffer

    raw = msg.payload.decode("utf-8", errors="replace").strip()

    # Try to parse as float, fall back to string
    try:
        value = float(raw)
    except ValueError:
        value = raw

    # Store in buffer by topic
    if msg.topic == "esp32/battery/voltage":
        reading_buffer["battery"] = value
    elif msg.topic == "esp32/solar/voltage":
        reading_buffer["solar"] = value

    # When both values arrived, write one row
    if reading_buffer["battery"] is not None and reading_buffer["solar"] is not None:
        flush_row(userdata)


def on_disconnect(client, userdata, disconnect_flags, rc, properties=None):
    print(f"[!] Disconnected (rc={rc})")


# -- Main -------------------------------------------------------
def main():
    global message_count

    wb, ws = get_or_create_workbook(EXCEL_FILE)
    print(f"[i] Logging to: {EXCEL_FILE}")

    userdata = {"wb": wb, "ws": ws}

    client = mqtt.Client(
        mqtt.CallbackAPIVersion.VERSION2,
        client_id=f"esp32_logger_{uuid.uuid4().hex[:6]}",
        userdata=userdata,
    )
    client.on_connect = on_connect
    client.on_message = on_message
    client.on_disconnect = on_disconnect

    # Graceful shutdown
    def shutdown(sig, frame):
        print(f"\n[i] Stopping... {message_count} rows logged.")
        client.disconnect()
        client.loop_stop()
        sys.exit(0)

    signal.signal(signal.SIGINT, shutdown)

    print(f"[>>] Connecting to {BROKER}:{PORT} ...")
    client.connect(BROKER, PORT, keepalive=60)
    client.loop_forever()


if __name__ == "__main__":
    main()
